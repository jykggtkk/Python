#!/usr/bin/python
# -*- coding: utf-8 -*-
'''
脚本功能：读取excel并初始化sqlite
输入参数：excel 
初次编写日期：2017-04-24
说明：依照excel配置初始化sqlite表结构，并插入各模板数据（插入待后续开发）
'''
 
import os
import sys
from openpyxl import load_workbook
import sqlite3
import logging

def main():

	#初始化建表
	logging.info( "Creating sqlite tables.")
	wb = load_workbook(filename='system_init.xlsx', read_only=True)
	ws = wb['meta_dict']
	
	#表清单-去重复
	tables=[]
	for i in range(2,ws.max_row+1):
		if ws.cell(column=1,row=i).value not in tables:
			tables.append(ws.cell(column=1,row=i).value)
		else:
			pass
	#每个表
 	for table in tables:
		cols=[]
		for row in ws.rows:
			if row[0].value == table:
				line = [col.value for col in row]
				cols.append(line)
			else:
				pass
		#获取语句
		logging.debug( "Table "+table+" init sql are producing...... ")
		sql=tab_sql(cols) 
		#执行SQL
		logging.debug( "Table "+table+" init sql are being executed...... ")
		tab_init(table,sql)
		logging.info( "Table "+table+" has been create. ")
	logging.info( "ALL sqlite tables have been create. ")

	#插入初始化数据


#组织sql
def tab_sql(cols):

	table=cols[0][0]
	#获取创建表语句，带主键
	txt=[]
	key=[]
	for col in cols:
		txt.append(str(col[1])+" "+str(col[2]))
		if col[6]==1:
			key.append(col[1])
		else:
			pass
	txts=','.join(txt)
	keys=','.join(key)
	#print table+","+keys
	#创建有主键和无主键
	if keys!="":
		sql ="create table "+table+"("+txts+","+ "CONSTRAINT AUTH_GROUP_PERMISSIONS_PK PRIMARY KEY ("+keys+"));"
	else:
		sql="create table "+table+"("+txts+");"
	return sql
    

def tab_init(table,sql):

	try:
		conn = sqlite3.connect('system.sqlite3')
		sqlpre="SELECT count(*) FROM sqlite_master WHERE type='table' AND name='"+table+"';"
		cursor = conn.execute(sqlpre)
		for row in cursor:
			flag=row[0]
		if flag==1:
			#sqlpre="create table bak_"+table+" as select * from "+table+";"    #先不考虑备份 只有表就删除
			logging.debug("Table "+table+" already exists in database,drop it first. ")
			sqlpre="drop table "+table+";\n"
			conn.execute(sqlpre)
		else:
			pass
		#创建新表
		conn.execute(sql)
		conn.commit()
		conn.close()
	except BaseException,err:
		logging.error("When creating table "+table+","+err.args)

if __name__ == '__main__': 
    #设置环境编码
    reload(sys)
    sys.setdefaultencoding('utf8')
    #初始化日志配置
    LOG_FORMAT = "%(asctime)s[%(levelname)s]%(message)s"
    DATE_FORMAT = '%Y-%m-%d %H:%M:%S' 
    logging.basicConfig(level=logging.INFO, format=LOG_FORMAT, datefmt=DATE_FORMAT)
    logging.info( "Beginning init. ")
    main()
