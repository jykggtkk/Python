#!/usr/bin/python
# -*- coding: utf-8 -*-
'''
脚本功能：读取excel并将数据插入到sqlite数据表中
输入参数：excel 
初次编写日期：2017-04-24
说明：依照excel 插入各模板数据 
'''
 
import os
import sys
from openpyxl import load_workbook
import sqlite3
import logging

def system_data_init(file,tab):
    #初始化建表
    wb = load_workbook(filename=file, data_only=True)
    ws = wb[tab]
    cols=get_cols(tab) 
    num=len(cols)
    cols_txt=','.join(cols)

    sql_txt="insert into "+tab+" ("+cols_txt+") values ("

    #获取语句
    logging.info( "Table "+tab+" data insert sql are producing...... ")
    #从第二行开始读取数据
    datas=[]
    for i in range(2,ws.max_row+1):
        data=[]
        for j in range(1,num+1):
            if ws.cell(column=j,row=i).value == None:
                val='Null'
            elif isinstance(ws.cell(column=j,row=i).value,unicode):
                val="'"+ws.cell(column=j,row=i).value.encode('utf-8')+"'"
            elif isinstance(ws.cell(column=j,row=i).value,str):
                val="'"+ws.cell(column=j,row=i).value+"'"
            else:
                val=str(ws.cell(column=j,row=i).value)
            data.append(val)
        data_txt=','.join(data) 
        sql=sql_txt+data_txt+");"
        datas.append(sql)

    #执行SQL
    try:
        conn=sqlite3.connect('system.sqlite3')
        delsql="delete from "+tab+";\n"
        cursor=conn.execute(delsql)
        #数据操作要提交事务
        conn.commit()
        logging.info( "Table "+tab+" has been truncate. ")
        for txt in datas:
            logging.debug( "Execute:"+txt)
            conn.execute(txt)
        conn.commit()
    except BaseException,err:
        logging.error("When execute sql "+txt+","+str(err.args))
    finally:
        conn.close()

    logging.info( "ALL data tables have been insert into "+tab+". ")

def get_cols(tab):
    try:
        cols=[]
        #tab='meta_dict'
        conn = sqlite3.connect('system.sqlite3')
        sqlpre="PRAGMA table_info( "+tab+");"
        cursor = conn.execute(sqlpre)
        for row in cursor:
            cols.append(str(row[1]))
        conn.close()
    except BaseException,err:
        logging.error("When creating table "+tab+","+err.args)
    return cols 
if __name__ == '__main__':
    #设置环境编码
    reload(sys)
    sys.setdefaultencoding('utf8')
    #初始化日志配置
    LOG_FORMAT = "%(asctime)s[%(levelname)s]%(message)s"
    DATE_FORMAT = '%Y-%m-%d %H:%M:%S' 
    logging.basicConfig(level=logging.DEBUG, format=LOG_FORMAT, datefmt=DATE_FORMAT)
    logging.info( "Beginning data init. ")

    #可改为输入参数
    system_data_init('c:\\git\\Python\\etlutil\\system_init.xlsx','meta_dict')


